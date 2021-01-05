import requests
import datetime
import random
import re

from openpyxl.styles import PatternFill
from openpyxl import Workbook
from openpyxl.styles import Font


def get_data_from_weather_api():
    url = "https://community-open-weather-map.p.rapidapi.com/weather"
    city = input("Give a city name: ")

    querystring = {"q": city, "units": "\"metric\""}

    headers = {

        'x-rapidapi-key': "here_goes_your_apikey",
        'x-rapidapi-host': "community-open-weather-map.p.rapidapi.com"
        }

    response = requests.request("GET", url, headers=headers, params=querystring)
    return response


def get_city_weather_info(response):
    response_json = response.json()

    print(response_json)

    info = {
        "city": response_json["name"],
        "weather": response_json["weather"][0]['description'],
        "temperature": int(response_json["main"]['temp'] - 273.15),
        "feels_like": int(response_json["main"]['feels_like'] - 273.15),
        "pressure": response_json["main"]['pressure'],
        "clouds": response_json["clouds"]['all'],
        "sunset": datetime.datetime.utcfromtimestamp(response_json["sys"]['sunset']).strftime('%H:%M'),
        "date": "",

    }

    UTC_now = datetime.datetime.utcnow()
    shift_in_s_from_UTC = response_json['timezone']

    info['date'] = (UTC_now + datetime.timedelta(0, shift_in_s_from_UTC)).strftime("%m/%d/%Y, %H:%M:%S")
    return info


def get_random_quote():
    url = "https://type.fit/api/quotes"

    response = requests.request("GET", url).json()
    return response[random.randint(0, len(response))]


def create_calendar_card(date, quote, all_data):
    workbook = Workbook()
    sheet = workbook.active

    # frame of the card
    for i in range(1, 20):
        for j in ["A", "B", "C", "D", "E", "F", "G", "H"]:
            sheet[str(j) + str(i)].fill = PatternFill(start_color="C2D9ED", fill_type="solid")

    # inside of the card - different color
    for i in range(3, 17):
        for j in ["B", "C", "D", "E", "F", "G"]:
            sheet[str(j) + str(i)].fill = PatternFill(start_color="edf1f5", fill_type="solid")

    # area for city name
    sheet.merge_cells('C5:E7', 'Merged Range')
    sheet['C5'] = all_data['city']
    sheet['C5'].font = Font(size="30")

    # area for date
    sheet.merge_cells('C8:E10', 'Merged Range')
    sheet['C8'] = date.split(",")[0]
    sheet['C8'].font = Font(size="28")

    # area for time
    sheet.merge_cells('C11:E13', 'Merged Range')
    sheet['C11'] = date.split(",")[1].strip()
    sheet['C11'].font = Font(size="28")

    # area for quote
    sheet.merge_cells('A2:G2', 'Merged Range2')
    sheet['A2'] = quote
    sheet['A2'].font = Font(size="10")

    # print weather info
    sheet['B14'] = 'temp: '
    sheet['C14'] = str(all_data['temperature']) + "°C"

    sheet['B15'] = 'feels like: '
    sheet['C15'] = str(all_data['feels_like']) + "°C"

    sheet['B16'] = 'pressure: '
    sheet['C16'] = all_data['pressure']

    sheet['E14'] = 'weather: '
    sheet['F14'] = all_data['weather']

    sheet['E15'] = all_data['clouds']
    sheet['F15'] = '% clouds'

    sheet['E16'] = 'sunset at: '
    sheet['F16'] = all_data['sunset']

    workbook.save(filename=all_data['city'] + re.sub("[/.:]", ".", str(all_data['date'])) + ".xlsx")


response = get_data_from_weather_api()

if response.status_code == 200:
    data = get_city_weather_info(response)
    create_calendar_card(date=data['date'], quote=get_random_quote()['text'], all_data=data)
elif response.status_code == 404:
    print("Check the spelling of city name.")
elif response.status_code == 500:
    print("Sorry, it's our fault. Come back later.")
